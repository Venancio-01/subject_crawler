import openpyxl


def init_excel():
    workbook = openpyxl.load_workbook('result.xlsx')
    sheet = workbook.active

    sheet.cell(row=1, column=1).value = '序号'
    sheet.cell(row=1, column=2).value = '题目'
    sheet.cell(row=1, column=3).value = '答案'

    sheet.delete_rows(2, sheet.max_row)
    return workbook, sheet


def edit_excel(sheet, subject_title, right_answer):
    for index, _ in enumerate(subject_title):
        sheet.cell(row=index + 2, column=1).value = index + 1
        sheet.cell(row=index + 2, column=2).value = subject_title[index] + '---' + right_answer[index]
        sheet.cell(row=index + 2, column=3).value = right_answer[index]


def save_excel(workbook):
    # 保存 Excel 文件
    workbook.save('result.xlsx')